import json
import re
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


BASE = Path(r"C:\Users\gabri\Documents\UTFPR\ArtigoMarcella")
OUT = BASE / "base_consolidada_saida"
OUT.mkdir(exist_ok=True)

MUN_TARGETS = {"Castro (PR)", "Ponta Grossa (PR)"}


def clean_text(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return str(x).replace("\xa0", " ").strip()


def slug(s):
    s = unicodedata.normalize("NFKD", clean_text(s)).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").lower()
    return s or "coluna"


def make_unique(names):
    seen = {}
    result = []
    for name in names:
        base = clean_text(name) or "coluna"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base}_{count}")
    return result


def parse_num(x):
    s = clean_text(x)
    if s in ["", "-", "...", "X", "x", "null", "NULL", "NaN", "nan"]:
        return np.nan
    if re.match(r"^-?\d{1,3}(\.\d{3})*,\d+$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan


def extract_year_unit(yraw):
    s = clean_text(yraw)
    m = re.search(r"(\d{4})", s)
    year = int(m.group(1)) if m else None
    unit = ""
    m2 = re.search(r"\((.*?)\)", s)
    if m2:
        unit = m2.group(1).strip()
    return year, unit


def detect_encoding(path):
    raw = path.read_bytes()[:4000]
    for enc in ["utf-8-sig", "latin1", "cp1252"]:
        try:
            raw.decode(enc)
            return enc
        except Exception:
            pass
    return "latin1"


def normalize_weather_datetime(df):
    cols_upper = {str(c).upper(): c for c in df.columns}
    date_candidates = [
        c for c in df.columns if "DATA" in str(c).upper() and "FUNDACAO" not in str(c).upper()
    ]
    time_candidates = [c for c in df.columns if "HORA" in str(c).upper()]
    date_col = date_candidates[0]
    time_col = time_candidates[0]
    date_text = df[date_col].astype(str).str.strip().str.replace("/", "-", regex=False)
    df["data"] = pd.to_datetime(date_text, errors="coerce")
    time_parts = df[time_col].astype(str).str.extract(r"(\d{1,2}):?(\d{2})")
    hour = time_parts[0].fillna("").str.zfill(2)
    minute = time_parts[1].fillna("").str.zfill(2)
    df["hora"] = hour + ":" + minute
    df["data_hora"] = pd.to_datetime(
        df["data"].dt.strftime("%Y-%m-%d") + " " + df["hora"], errors="coerce"
    )
    df["ano"] = df["data_hora"].dt.year
    return date_col, time_col


def parse_weather_file(path, source_kind):
    enc = detect_encoding(path)
    lines = path.read_text(encoding=enc, errors="replace").splitlines()
    header_idx = None
    meta = {}
    for i, line in enumerate(lines):
        up = line.upper()
        if ";" in line and "DATA" in up and "HORA" in up:
            header_idx = i
            break
        if ";" in line:
            k, v = line.split(";", 1)
            meta[clean_text(k).rstrip(":")] = clean_text(v)
        elif ":" in line:
            k, v = line.split(":", 1)
            meta[clean_text(k).rstrip(":")] = clean_text(v)
    if header_idx is None:
        raise ValueError(f"Header not found in {path.name}")

    df = pd.read_csv(path, sep=";", skiprows=header_idx, encoding=enc, dtype=str, engine="python")
    df = df.dropna(axis=1, how="all")
    df = df.loc[:, [c for c in df.columns if not str(c).startswith("Unnamed")]]
    df.columns = make_unique([clean_text(c) for c in df.columns])
    date_col, time_col = normalize_weather_datetime(df)

    fixed_cols = ["data", "hora", "data_hora", "ano"]
    rename = {}
    for c in list(df.columns):
        if c in ["data", "hora", "data_hora", "ano"]:
            continue
        new = slug(c)
        base_new = new
        k = 2
        while new in fixed_cols:
            new = f"{base_new}_{k}"
            k += 1
        rename[c] = new
        fixed_cols.append(new)
    df = df.rename(columns=rename)

    date_slug = rename.get(date_col, slug(date_col))
    time_slug = rename.get(time_col, slug(time_col))
    for c in fixed_cols:
        if c in [date_slug, time_slug]:
            continue
        s = df[c].astype(str).str.strip()
        s = s.replace({"": np.nan, "null": np.nan, "NULL": np.nan, "nan": np.nan})
        conv = pd.to_numeric(s.str.replace(",", ".", regex=False), errors="coerce")
        if conv.notna().sum() > 0:
            df[c] = conv
        else:
            df[c] = s.replace({"nan": np.nan})

    df["arquivo_origem"] = path.name
    df["fonte_clima"] = source_kind
    for k, v in meta.items():
        df[f"meta_{slug(k)}"] = v

    return df, {
        "arquivo": path.name,
        "fonte": source_kind,
        "encoding": enc,
        "header_linha": header_idx + 1,
        "linhas_lidas": len(df),
        "data_min": str(df["data_hora"].min()),
        "data_max": str(df["data_hora"].max()),
        "anos": ",".join(map(str, sorted([int(x) for x in df["ano"].dropna().unique()]))),
    }


def extract_productivity():
    prod_rows = []
    notes_rows = []
    prod_audit = []
    xlsx_files = sorted(BASE.glob("*.xlsx"))
    for p in xlsx_files:
        wb = load_workbook(p, read_only=True, data_only=True)
        for ws in wb.worksheets:
            sheet = ws.title
            max_row, max_col = ws.max_row, ws.max_column
            if sheet.lower().startswith("notas"):
                extracted = 0
                for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    vals = [clean_text(v) for v in row]
                    if any(vals):
                        notes_rows.append(
                            {
                                "arquivo": p.name,
                                "aba": sheet,
                                "linha": r,
                                "texto": " | ".join([v for v in vals if v]),
                            }
                        )
                        extracted += 1
                prod_audit.append(
                    {
                        "arquivo": p.name,
                        "aba": sheet,
                        "tipo": "notas",
                        "linhas_lidas": max_row,
                        "colunas_lidas": max_col,
                        "linhas_extraidas": extracted,
                    }
                )
                continue

            rows = list(ws.iter_rows(values_only=True))
            titulo = clean_text(rows[0][0] if len(rows) > 0 and rows[0] else "")
            variavel = clean_text(rows[1][0] if len(rows) > 1 and rows[1] else "")
            variavel = re.sub(r"^Vari[aá]vel\s*-\s*", "", variavel, flags=re.I)

            current_year_raw = ""
            col_meta = {}
            year_row = rows[3] if len(rows) > 3 else []
            product_row = rows[4] if len(rows) > 4 else []
            for idx in range(3, max_col):
                yr_cell = clean_text(year_row[idx] if idx < len(year_row) else "")
                if yr_cell:
                    current_year_raw = yr_cell
                prod = clean_text(product_row[idx] if idx < len(product_row) else "")
                year, unit_from_year = extract_year_unit(current_year_raw)
                if year and prod:
                    col_meta[idx] = {
                        "ano": year,
                        "ano_cabecalho": current_year_raw,
                        "produto": prod,
                        "unidade_cabecalho": unit_from_year,
                    }

            extracted = 0
            for r_idx, row in enumerate(rows[5:], start=6):
                municipio = clean_text(row[2] if len(row) > 2 else "")
                if municipio not in MUN_TARGETS:
                    continue
                nivel = clean_text(row[0] if len(row) > 0 else "")
                cod = clean_text(row[1] if len(row) > 1 else "")
                for c_idx, meta in col_meta.items():
                    raw = clean_text(row[c_idx] if c_idx < len(row) else "")
                    prod_rows.append(
                        {
                            "tipo_registro": "produtividade",
                            "arquivo_origem": p.name,
                            "aba_origem": sheet,
                            "titulo_tabela": titulo,
                            "variavel": variavel,
                            "nivel": nivel,
                            "codigo_municipio": cod,
                            "municipio": municipio,
                            "ano": meta["ano"],
                            "produto": meta["produto"],
                            "unidade_cabecalho": meta["unidade_cabecalho"],
                            "valor_original": raw,
                            "valor_numerico": parse_num(raw),
                            "linha_origem": r_idx,
                            "coluna_origem": c_idx + 1,
                            "ano_cabecalho_original": meta["ano_cabecalho"],
                        }
                    )
                    extracted += 1
            prod_audit.append(
                {
                    "arquivo": p.name,
                    "aba": sheet,
                    "tipo": "produtividade",
                    "linhas_lidas": max_row,
                    "colunas_lidas": max_col,
                    "colunas_ano_produto": len(col_meta),
                    "linhas_extraidas": extracted,
                }
            )
        wb.close()
    return pd.DataFrame(prod_rows), pd.DataFrame(notes_rows), pd.DataFrame(prod_audit), xlsx_files


def main():
    prod_df, notes_df, prod_audit_df, xlsx_files = extract_productivity()

    annual_files = sorted(BASE.glob("INMET_S_PR_A819_CASTRO_*.CSV"))
    hist_file = BASE / "dados_A819_H_2006-07-08_2026-01-01.csv"
    weather_parts = []
    weather_audit = []
    for p in annual_files:
        df, audit = parse_weather_file(p, "csv_anual_inmet")
        weather_parts.append(df)
        weather_audit.append(audit)

    annual_weather = pd.concat(weather_parts, ignore_index=True, sort=False)
    hist_weather, hist_audit = parse_weather_file(hist_file, "csv_historico_complementar")
    weather_audit.append(hist_audit)

    annual_key_series = pd.to_datetime(annual_weather["data_hora"], errors="coerce")
    hist_key_series = pd.to_datetime(hist_weather["data_hora"], errors="coerce")
    annual_keys = set(annual_key_series.dropna().astype("int64").tolist())
    hist_missing = hist_weather[~hist_key_series.astype("int64").isin(annual_keys)].copy()
    hist_missing["fonte_clima"] = "csv_historico_usado_para_lacunas"
    weather_df = pd.concat([annual_weather, hist_missing], ignore_index=True, sort=False)
    weather_df = weather_df.sort_values("data_hora").reset_index(drop=True)
    weather_df["tipo_registro"] = "clima_castro_horario"
    weather_audit_df = pd.DataFrame(weather_audit)

    prod_for_union = prod_df.copy()
    weather_for_union = weather_df.copy()
    prod_for_union["data_hora"] = pd.NaT
    prod_for_union["data"] = ""
    prod_for_union["hora"] = ""
    prod_for_union["fonte_clima"] = ""
    weather_for_union["municipio"] = "Castro (PR)"
    weather_for_union["codigo_municipio"] = ""
    weather_for_union["produto"] = ""
    weather_for_union["variavel"] = ""
    weather_for_union["valor_original"] = ""
    weather_for_union["valor_numerico"] = np.nan
    weather_for_union["aba_origem"] = ""
    weather_for_union["titulo_tabela"] = ""
    weather_for_union["nivel"] = ""
    weather_for_union["unidade_cabecalho"] = ""
    weather_for_union["linha_origem"] = ""
    weather_for_union["coluna_origem"] = ""
    weather_for_union["ano_cabecalho_original"] = ""

    common_first = [
        "tipo_registro",
        "arquivo_origem",
        "aba_origem",
        "fonte_clima",
        "data_hora",
        "data",
        "hora",
        "ano",
        "municipio",
        "codigo_municipio",
        "nivel",
        "variavel",
        "produto",
        "unidade_cabecalho",
        "valor_original",
        "valor_numerico",
        "titulo_tabela",
        "linha_origem",
        "coluna_origem",
        "ano_cabecalho_original",
    ]
    all_cols = []
    for c in common_first + list(prod_for_union.columns) + list(weather_for_union.columns):
        if c not in all_cols:
            all_cols.append(c)

    base_df = pd.concat(
        [prod_for_union.reindex(columns=all_cols), weather_for_union.reindex(columns=all_cols)],
        ignore_index=True,
        sort=False,
    )

    for df in [weather_df, base_df]:
        if "data_hora" in df.columns:
            df["data_hora"] = pd.to_datetime(df["data_hora"], errors="coerce").dt.strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        if "data" in df.columns and not pd.api.types.is_string_dtype(df["data"]):
            df["data"] = pd.to_datetime(df["data"], errors="coerce").dt.strftime("%Y-%m-%d")

    prod_path = OUT / "produtividade_castro_ponta_grossa_longa.csv"
    weather_path = OUT / "clima_castro_horario.csv"
    base_path = OUT / "base_dados_consolidada_longa.csv"
    notes_path = OUT / "notas_produtividade.csv"
    audit_path = OUT / "auditoria_fontes.csv"
    summary_path = OUT / "resumo_consolidacao.json"

    prod_df.to_csv(prod_path, index=False, encoding="utf-8-sig")
    weather_df.to_csv(weather_path, index=False, encoding="utf-8-sig")
    base_df.to_csv(base_path, index=False, encoding="utf-8-sig")
    notes_df.to_csv(notes_path, index=False, encoding="utf-8-sig")

    audit_all = pd.concat(
        [prod_audit_df.assign(grupo="produtividade"), weather_audit_df.assign(grupo="clima")],
        ignore_index=True,
        sort=False,
    )
    audit_all.to_csv(audit_path, index=False, encoding="utf-8-sig")

    summary = {
        "arquivos_excel_lidos": len(xlsx_files),
        "arquivos_clima_anuais_lidos": len(annual_files),
        "arquivo_historico_complementar": hist_file.name,
        "linhas_produtividade": int(len(prod_df)),
        "linhas_clima_anuais": int(len(annual_weather)),
        "linhas_clima_complementadas_do_historico": int(len(hist_missing)),
        "linhas_clima_final": int(len(weather_df)),
        "linhas_base_consolidada": int(len(base_df)),
        "municipios_produtividade": sorted(prod_df["municipio"].dropna().unique().tolist()),
        "anos_produtividade": [int(x) for x in sorted(prod_df["ano"].dropna().unique())],
        "anos_clima_final": [
            int(x) for x in sorted(pd.to_numeric(weather_df["ano"], errors="coerce").dropna().unique())
        ],
        "data_hora_clima_min": str(pd.to_datetime(weather_df["data_hora"]).min()),
        "data_hora_clima_max": str(pd.to_datetime(weather_df["data_hora"]).max()),
        "saidas": {
            "produtividade": str(prod_path),
            "clima": str(weather_path),
            "base": str(base_path),
            "notas": str(notes_path),
            "auditoria": str(audit_path),
        },
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
